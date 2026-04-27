import hashlib
import json
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook


def normalize_grid(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return []
    m = max(len(r) for r in rows)
    return [r + [""] * (m - len(r)) for r in rows]


def xlsx_first_sheet_to_grid(blob: bytes) -> list[list[str]]:
    wb = load_workbook(BytesIO(blob), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in row])
    finally:
        wb.close()
    return normalize_grid(rows)


def grid_to_xlsx_bytes(grid: list[list[str]], sheet_title: str = "Sheet1") -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_title[:31] or "Sheet1"
    for r, row in enumerate(grid, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def stable_grid_hash(grid: list[list[str]]) -> str:
    payload = json.dumps(grid, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def parse_grid_form(cells: list[str] | None, nrows: int, ncols: int) -> list[list[str]]:
    """Rebuild grid from flattened form field `cell_0_0`, ... or indexed list."""
    if not cells:
        return []
    grid: list[list[str]] = []
    if isinstance(cells, str):
        cells = [cells]
    # htmx may send cell_r_c as separate keys — handled in router
    return [["" for _ in range(ncols)] for _ in range(nrows)]


def grid_from_flat_map(flat: dict[str, str], nrows: int, ncols: int) -> list[list[str]]:
    grid = [[""] * ncols for _ in range(nrows)]
    for key, val in flat.items():
        if not key.startswith("cell_"):
            continue
        try:
            _, r_s, c_s = key.split("_", 2)
            r, c = int(r_s), int(c_s)
            if 0 <= r < nrows and 0 <= c < ncols:
                grid[r][c] = val
        except (ValueError, IndexError):
            continue
    return grid
