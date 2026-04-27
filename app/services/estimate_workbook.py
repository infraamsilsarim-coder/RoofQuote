from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


RED_TITLE = Font(color="00FF0000", bold=True, size=16)
BLACK_BOLD = Font(bold=True)
UNDERLINE_BOLD = Font(bold=True, underline="single")
HEADING_DEF = Font(bold=True, size=13)


def _merge_title_row(ws, row: int, text: str, col_span: int = 8) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = RED_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    return row + 1


def write_metadata(
    ws,
    start_row: int,
    *,
    display_code: str,
    prepared_by: str,
    address: str,
    col_span: int = 8,
) -> int:
    r = start_row
    ws.cell(row=r, column=1, value=display_code or "PROJECT")
    ws.cell(row=r, column=1).font = Font(bold=True, size=18)
    r += 1
    r = _merge_title_row(ws, r, "Ranger Roofing and Solar", col_span)
    d = datetime.now()
    ws.cell(row=r, column=1, value="Prepared on:")
    ws.cell(row=r, column=2, value=f"{d.month}/{d.day}/{d.year}")
    r += 1
    ws.cell(row=r, column=1, value="Prepared by:")
    ws.cell(row=r, column=2, value=prepared_by)
    r += 1
    ws.cell(row=r, column=1, value="Address:")
    ws.cell(row=r, column=2, value=address)
    r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_span)
    c = ws.cell(row=r, column=1, value="BREAKDOWN")
    c.font = UNDERLINE_BOLD
    c.alignment = Alignment(horizontal="center")
    r += 1
    r += 1
    return r


def _cell_str(val: Any) -> Any:
    if val is None:
        return ""
    return val


def _append_labor_table(ws, r: int, rows: list[dict[str, Any]]) -> int:
    headers = ["Type", "Hrs/Day", "Rate", "Workers", "Days", "Total"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r, column=i, value=h).font = BLACK_BOLD
    r += 1
    for row in rows:
        workers = row.get("workers", row.get("num_workers", ""))
        ws.cell(row=r, column=1, value=row.get("type", ""))
        ws.cell(row=r, column=2, value=row.get("hours_per_day", ""))
        ws.cell(row=r, column=3, value=row.get("rate", ""))
        ws.cell(row=r, column=4, value=workers)
        ws.cell(row=r, column=5, value=row.get("days", ""))
        ws.cell(row=r, column=6, value=row.get("total", ""))
        r += 1
    return r


def _append_materials_table(ws, r: int, materials: list[dict[str, Any]]) -> int:
    headers = ["Item", "Unit", "Qty", "Price", "Total"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r, column=i, value=h).font = BLACK_BOLD
    r += 1
    for m in materials:
        ws.cell(row=r, column=1, value=m.get("item", ""))
        ws.cell(row=r, column=2, value=_cell_str(m.get("unit")))
        ws.cell(row=r, column=3, value=m.get("quantity", ""))
        ws.cell(row=r, column=4, value=m.get("price", ""))
        ws.cell(row=r, column=5, value=m.get("total", ""))
        r += 1
    return r


def _is_compact_deficiency_schema(obj: dict[str, Any]) -> bool:
    """New prompt.txt shape: deficiency_name + pricing_breakdown.travel_and_labor, no legacy status."""
    if not isinstance(obj, dict):
        return False
    if obj.get("status") in (
        "no_deficiency",
        "image_unclear",
        "deficiency_found",
        "model_error",
    ):
        return False
    if "deficiency_name" not in obj:
        return False
    pb = obj.get("pricing_breakdown")
    if not isinstance(pb, dict):
        return False
    return isinstance(pb.get("travel_and_labor"), list)


def _write_compact_deficiency(
    ws, r: int, obj: dict[str, Any], deficiency_number: int
) -> int:
    """Write one LLM JSON block: heading Deficiency #N, then new schema fields."""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1, value=f"Deficiency # {deficiency_number}")
    c.font = HEADING_DEF
    r += 1

    name = obj.get("deficiency_name", "")
    ws.cell(row=r, column=1, value="Deficiency:").font = BLACK_BOLD
    ws.cell(row=r, column=2, value=name)
    r += 1

    rates = obj.get("rates") or {}
    if rates:
        ws.cell(row=r, column=1, value="Rates").font = BLACK_BOLD
        r += 1
        ws.cell(row=r, column=1, value="Labor $/hr:")
        ws.cell(row=r, column=2, value=rates.get("labor_per_hour", ""))
        r += 1
        ws.cell(row=r, column=1, value="Travel $/hr:")
        ws.cell(row=r, column=2, value=rates.get("travel_rate", ""))
        r += 1

    pb = obj.get("pricing_breakdown") or {}
    labor_rows = pb.get("travel_and_labor") or []
    if labor_rows:
        ws.cell(row=r, column=1, value="Travel & labor").font = BLACK_BOLD
        r += 1
        r = _append_labor_table(ws, r, labor_rows)
    ws.cell(row=r, column=1, value="Total travel & labor:").font = BLACK_BOLD
    ws.cell(row=r, column=2, value=pb.get("total_travel_and_labor", ""))
    r += 1

    mats = obj.get("materials") or []
    if mats:
        ws.cell(row=r, column=1, value="Materials").font = BLACK_BOLD
        r += 1
        r = _append_materials_table(ws, r, mats)

    fs = obj.get("financial_summary") or {}
    if fs:
        ws.cell(row=r, column=1, value="Financial summary").font = BLACK_BOLD
        r += 1
        for label, key in (
            ("Subtotal materials", "subtotal_materials"),
            ("Tax rate", "tax_rate"),
            ("Tax amount", "tax_amount"),
            ("Total materials", "total_materials"),
            ("Deficiency price / unit", "deficiency_price_per_unit"),
            ("Number of units", "number_of_units"),
            ("Deficiency total", "deficiency_total"),
            ("Grand total", "grand_total"),
        ):
            if key in fs:
                ws.cell(row=r, column=1, value=f"{label}:")
                ws.cell(row=r, column=2, value=fs.get(key))
                r += 1

    sow = obj.get("scope_of_work")
    if sow:
        ws.cell(row=r, column=1, value="Scope of work").font = BLACK_BOLD
        r += 1
        if isinstance(sow, list):
            for i, step in enumerate(sow, start=1):
                ws.cell(row=r, column=1, value=f"{i}. {step}")
                r += 1
        elif isinstance(sow, dict):
            if sow.get("method"):
                ws.cell(row=r, column=1, value=str(sow["method"]))
                r += 1
            for i, step in enumerate(sow.get("steps") or [], start=1):
                ws.cell(row=r, column=1, value=f"{i}. {step}")
                r += 1

    r += 1
    return r


def _write_deficiency_detail(ws, r: int, d: dict[str, Any]) -> int:
    num = d.get("deficiency_number", "")
    desc = d.get("description", "")
    ws.cell(row=r, column=1, value=f"Deficiency #{num}: {desc}").font = BLACK_BOLD
    r += 1
    det = d.get("affected_quantity") or {}
    if det:
        ws.cell(row=r, column=1, value="Affected quantity:")
        ws.cell(row=r, column=2, value=f"{det.get('value', '')} {det.get('unit', '')}")
        r += 1
        ws.cell(row=r, column=1, value="Estimation:")
        ws.cell(row=r, column=2, value=str(det.get("estimation_method", "")))
        r += 1
    pb = d.get("pricing_breakdown") or {}
    ws.cell(row=r, column=1, value="Wage type:").font = BLACK_BOLD
    ws.cell(row=r, column=2, value=str(pb.get("wage_type", "")))
    r += 1
    labor = pb.get("labor_and_travel") or []
    if labor:
        ws.cell(row=r, column=1, value="Labor & Travel").font = BLACK_BOLD
        r += 1
        r = _append_labor_table(ws, r, labor)
    ws.cell(row=r, column=1, value="Total travel & labor:")
    ws.cell(row=r, column=2, value=pb.get("total_travel_and_labor", ""))
    r += 1
    lb = pb.get("labor_burden") or {}
    if lb:
        ws.cell(row=r, column=1, value="Labor burden:")
        ws.cell(
            row=r,
            column=2,
            value=f"{lb.get('rate', '')} → {lb.get('amount', '')}",
        )
        r += 1
    mats = pb.get("materials") or []
    if mats:
        ws.cell(row=r, column=1, value="Materials").font = BLACK_BOLD
        r += 1
        r = _append_materials_table(ws, r, mats)
    ws.cell(row=r, column=1, value="Material subtotal:")
    ws.cell(row=r, column=2, value=pb.get("material_subtotal", ""))
    r += 1
    ws.cell(row=r, column=1, value="Tax:")
    ws.cell(
        row=r,
        column=2,
        value=f"{pb.get('tax_rate', '')} → {pb.get('tax_amount', '')}",
    )
    r += 1
    ws.cell(row=r, column=1, value="Total materials (w/ tax):")
    ws.cell(row=r, column=2, value=pb.get("total_materials", ""))
    r += 1
    tot = d.get("totals") or {}
    if tot:
        ws.cell(row=r, column=1, value="Deficiency total:").font = BLACK_BOLD
        ws.cell(row=r, column=2, value=tot.get("deficiency_total", ""))
        r += 1
    sow = d.get("scope_of_work") or {}
    if sow:
        ws.cell(row=r, column=1, value="Scope of work").font = BLACK_BOLD
        r += 1
        if isinstance(sow, list):
            for i, step in enumerate(sow, start=1):
                ws.cell(row=r, column=1, value=f"{i}. {step}")
                r += 1
        elif isinstance(sow, dict):
            if sow.get("method"):
                ws.cell(row=r, column=1, value=str(sow["method"]))
                r += 1
            for i, step in enumerate(sow.get("steps") or [], start=1):
                ws.cell(row=r, column=1, value=f"{i}. {step}")
                r += 1
    r += 1
    return r


def _write_image_result_block(ws, r: int, idx: int, obj: dict[str, Any]) -> int:
    # Do not start cell text with "=" — Excel treats it as a formula ("=== …" → #ERROR!).
    ws.cell(row=r, column=1, value=f"Site photo #{idx + 1} — analysis").font = Font(
        bold=True, size=12
    )
    r += 1
    status = str(obj.get("status", "") or "")

    if status in ("no_deficiency", "image_unclear"):
        ws.cell(row=r, column=1, value="Notes:")
        ws.cell(row=r, column=2, value=str(obj.get("notes", "")))
        r += 2
        return r

    if status == "model_error":
        ws.cell(row=r, column=1, value="Error:")
        ws.cell(row=r, column=2, value=str(obj.get("notes", "")))
        r += 1
        if obj.get("photo_filename"):
            ws.cell(row=r, column=1, value="Photo:")
            ws.cell(row=r, column=2, value=str(obj.get("photo_filename", "")))
            r += 1
        r += 1
        return r

    if status != "deficiency_found":
        ws.cell(row=r, column=1, value="Raw / unrecognized response")
        r += 1
        return r

    ia = obj.get("image_analysis") or {}
    if ia.get("description"):
        ws.cell(row=r, column=1, value="Image description:").font = BLACK_BOLD
        r += 1
        ws.cell(row=r, column=1, value=str(ia["description"]))
        r += 1

    if "deficiencies" in obj and isinstance(obj["deficiencies"], list):
        for d in obj["deficiencies"]:
            r = _write_deficiency_detail(ws, r, d)
        pt = obj.get("project_totals") or {}
        if pt:
            ws.cell(row=r, column=1, value="Project totals (this image)").font = BLACK_BOLD
            r += 1
            for k, v in pt.items():
                ws.cell(row=r, column=1, value=k)
                ws.cell(row=r, column=2, value=v)
                r += 1
            r += 1
    elif "deficiency_details" in obj:
        r = _write_deficiency_detail(ws, r, obj["deficiency_details"])
    r += 1
    return r


def build_estimate_workbook(
    *,
    display_code: str,
    prepared_by: str,
    address: str,
    results: list[dict[str, Any]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Estimate"
    r = write_metadata(ws, 1, display_code=display_code, prepared_by=prepared_by, address=address)

    deficiency_index = 0
    for i, obj in enumerate(results):
        if not isinstance(obj, dict):
            continue
        if _is_compact_deficiency_schema(obj):
            deficiency_index += 1
            r = _write_compact_deficiency(ws, r, obj, deficiency_index)
        else:
            r = _write_image_result_block(ws, r, i, obj)

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
