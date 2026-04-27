from io import BytesIO

from openpyxl import load_workbook


def xlsx_bytes_as_text(blob: bytes, max_chars: int = 900_000) -> str:
    """Extract tab-separated sheet text from .xlsx bytes for LLM context."""
    wb = load_workbook(BytesIO(blob), read_only=True, data_only=True)
    chunks: list[str] = []
    try:
        for sheet in wb.worksheets:
            lines: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                lines.append(
                    "\t".join("" if c is None else str(c) for c in row).rstrip()
                )
            body = "\n".join(lines).strip()
            if body:
                chunks.append(f"### Sheet: {sheet.title}\n{body}")
    finally:
        wb.close()
    text = "\n\n".join(chunks)
    if len(text) > max_chars:
        return text[:max_chars] + "\n\n[... truncated for size ...]"
    return text
