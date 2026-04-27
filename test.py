import base64
import json
import os
import sys

import requests
from openpyxl import load_workbook

# ── Config ──────────────────────────────────────────────
OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY", "")
MODEL = os.environ.get("OPENROUTER_MODEL", "anthropic/claude-opus-4.6")
API_URL = "https://openrouter.ai/api/v1/chat/completions"

# OpenRouter chat completions: PDFs and other files use type "file", not Anthropic "document".
# See https://openrouter.ai/docs/guides/overview/multimodal/pdfs

# ── Load system prompt ──────────────────────────────────
SYSTEM_PROMPT = open("prompt.txt", encoding="utf-8").read()


def to_b64(path: str) -> str:
    with open(path, "rb") as f:
        return base64.standard_b64encode(f.read()).decode("utf-8")


def xlsx_as_text(path: str, max_chars: int = 900_000) -> str:
    """OpenRouter rejects .xlsx as file uploads; send extracted values as text."""
    wb = load_workbook(path, read_only=True, data_only=True)
    chunks: list[str] = []
    try:
        for sheet in wb.worksheets:
            lines: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                lines.append(
                    "\t".join("" if c is None else str(c) for c in row).rstrip()
                )
            body = "\n".join(lines).strip()
            if body:
                chunks.append(f"### Sheet: {sheet.title}\n{body}")
    finally:
        wb.close()
    text = "\n\n".join(chunks)
    if len(text) > max_chars:
        return text[:max_chars] + "\n\n[... truncated for size ...]"
    return text


def main() -> None:
    if not OPENROUTER_API_KEY:
        print("Set OPENROUTER_API_KEY in your environment.", file=sys.stderr)
        sys.exit(1)

    # ── Load inputs ─────────────────────────────────────────
    iroof_b64 = to_b64("iroof_results.pdf")
    pricing_text = xlsx_as_text("Ranger_Roofing_Master_Pricing_v1.xlsx")
    photo_b64 = to_b64("Screenshot 2026-04-27 191745.png")

    notes = "Prevailing wage project. Sacramento site."

    user_content = [
        {
            "type": "file",
            "file": {
                "filename": "iroof_results.pdf",
                "file_data": f"data:application/pdf;base64,{iroof_b64}",
            },
        },
        {
            "type": "text",
            "text": (
                "## Master pricing workbook (from Ranger_Roofing_Master_Pricing_v1.xlsx)\n"
                "Tab-separated rows per sheet. Use only these rates and SKUs.\n\n"
                f"{pricing_text}"
            ),
        },
        {
            "type": "image_url",
            "image_url": {
                "url": f"data:image/png;base64,{photo_b64}",
            },
        },
        {
            "type": "text",
            "text": (
                "Analyze this roof photo for deficiencies.\n\n"
                f"NOTES: {notes}"
            ),
        },
    ]

    response = requests.post(
        API_URL,
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://rangerroofing.com",
            "X-Title": "RoofQuote AI",
        },
        json={
            "model": MODEL,
            "max_tokens": 4096,
            "messages": [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_content},
            ],
        },
    )

    data = response.json()

    if "error" in data:
        print(f"API Error: {data['error']}")
        return

    result_text = data["choices"][0]["message"]["content"]

    if "```json" in result_text:
        json_str = result_text.split("```json")[1].split("```")[0].strip()
    else:
        json_str = result_text.strip()

    try:
        result = json.loads(json_str)
        print(json.dumps(result, indent=2))
    except json.JSONDecodeError:
        print(result_text)


if __name__ == "__main__":
    main()
